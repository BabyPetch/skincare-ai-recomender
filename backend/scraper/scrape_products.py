from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
import pandas as pd
import time
import random
import os
import datetime

# ===============================
# CONFIG
# ===============================

HEADLESS = True
INPUT_FILE = "product_links.txt"
AUTOSAVE_EVERY = 50

timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = f"products_{timestamp}.xlsx"

# ===============================
# TAG RULES
# ===============================

ACTIVE_RULES = {
    "retinol":          ["retinol"],
    "retinal":          ["retinal"],
    "aha":              ["glycolic acid", "lactic acid", "mandelic acid"],
    "bha":              ["salicylic acid"],
    "pha":              ["gluconolactone"],
    "vitamin_c":        ["ascorb", "ascorbyl"],
    "niacinamide":      ["niacinamide"],
    "peptide":          ["peptide"],
    "ceramide":         ["ceramide"],
    "zinc_oxide":       ["zinc oxide"],
    "titanium_dioxide": ["titanium dioxide"],
}

FUNCTION_RULES = {
    "brightening":    ["ascorb", "niacinamide", "arbutin"],
    "anti_aging":     ["retinol", "retinal", "peptide"],
    "acne_control":   ["salicylic acid"],
    "calming":        ["centella", "allantoin"],
    "barrier_repair": ["ceramide"],
    "hydrating":      ["hyaluronic", "glycerin"],
    "exfoliating":    ["glycolic", "salicylic", "lactic"],
}

# ===============================
# AUTO-DETECT CATEGORY
# ===============================

CATEGORY_FROM_URL = {
    "sunscreen":    ["sunscreen", "spf", "sun-protect", "uv-"],
    "serum":        ["serum", "essence", "ampoule", "booster"],
    "moisturizer":  ["cream", "moisturizer", "lotion", "balm", "butter"],
    "mask":         ["mask", "sheet-mask", "pack"],
    "cleanser":     ["cleanser", "wash", "foam", "cleansing", "micellar"],
    "toner":        ["toner", "mist", "softener"],
    "eye_care":     ["eye-cream", "eye-serum", "eye-gel"],
    "lip_care":     ["lip-balm", "lip-mask", "lip-care"],
    "exfoliator":   ["exfoliant", "peeling", "scrub"],
    "oil":          ["-oil", "facial-oil", "face-oil"],
}

CATEGORY_FROM_NAME = {
    "sunscreen":    ["sunscreen", "spf", "sun protect", "uv", "solar"],
    "serum":        ["serum", "essence", "ampoule", "booster", "concentrate"],
    "moisturizer":  ["cream", "moisturizer", "lotion", "balm", "emulsion"],
    "mask":         ["mask", "pack", "sheet mask"],
    "cleanser":     ["cleanser", "wash", "foam", "cleansing oil", "micellar"],
    "toner":        ["toner", "tonic", "mist", "softener", "essence toner"],
    "eye_care":     ["eye cream", "eye serum", "eye gel", "eye treatment"],
    "lip_care":     ["lip balm", "lip mask", "lip care", "lip treatment"],
    "exfoliator":   ["exfoliant", "exfoliator", "peeling", "scrub", "peel"],
    "oil":          ["facial oil", "face oil", "dry oil"],
}

CATEGORY_FROM_INGREDIENTS = {
    "sunscreen":   [["zinc oxide", "titanium dioxide", "avobenzone",
                        "octinoxate", "octisalate", "octocrylene"]],
    "exfoliator":  [["glycolic acid", "salicylic acid", "lactic acid",
                        "mandelic acid", "gluconolactone"]],
    "moisturizer": [["ceramide", "hyaluronic acid", "squalane",
                        "shea butter", "glycerin"]],
}

def guess_category(url, name, ingredients_raw):
    url_lower  = url.lower()
    name_lower = name.lower()
    ingr_lower = ingredients_raw.lower()

    # 1) จาก URL
    for cat, keywords in CATEGORY_FROM_URL.items():
        if any(kw in url_lower for kw in keywords):
            return cat

    # 2) จากชื่อสินค้า
    for cat, keywords in CATEGORY_FROM_NAME.items():
        if any(kw in name_lower for kw in keywords):
            return cat

    # 3) จาก ingredients (majority vote)
    scores = {}
    for cat, kw_list in CATEGORY_FROM_INGREDIENTS.items():
        flat = kw_list[0]
        hit = sum(1 for kw in flat if kw in ingr_lower)
        if hit:
            scores[cat] = hit
    if scores:
        return max(scores, key=scores.get)

    return "unknown"

# ===============================
# HELPER FUNCTIONS
# ===============================

def clean_text(text):
    return " ".join(text.split())

def detect_subtype(name, ingredients, major_category):
    text = (name + " " + ingredients).lower()
    if major_category == "sunscreen":
        if "zinc oxide" in text or "titanium dioxide" in text:
            if "avobenzone" in text or "octinoxate" in text or "octocrylene" in text:
                return "hybrid_sunscreen"
            return "mineral_sunscreen"
        return "chemical_sunscreen"
    if "retinol" in text:
        return "retinol_serum"
    if "retinal" in text:
        return "retinal_serum"
    if "salicylic acid" in text:
        return "bha_exfoliator"
    if "glycolic acid" in text:
        return "aha_exfoliator"
    if "lactic acid" in text:
        return "aha_exfoliator"
    return major_category

def detect_tags(text, rules):
    tags, text = set(), text.lower()
    for tag, keywords in rules.items():
        if any(kw in text for kw in keywords):
            tags.add(tag)
    return sorted(tags)

# ===============================
# EXCEL BUILDER
# ===============================

# Color palette
NAVY      = "1B3A5C"
TEAL      = "2D7D9A"
LIGHT_BG  = "EBF4F8"
WHITE     = "FFFFFF"
ACCENT1   = "E8F5E9"   # green tint
ACCENT2   = "FFF3E0"   # orange tint
ACCENT3   = "F3E5F5"   # purple tint
GRAY_LINE = "D0D7DE"

def make_border(color=GRAY_LINE):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row, col, value, bg=NAVY, fg=WHITE, size=11, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=wrap)
    c.border    = make_border()
    return c

def data_cell(ws, row, col, value, bg=WHITE, bold=False, wrap=False,
              align="left", color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=10, bold=bold, color=color)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap)
    c.border    = make_border()
    return c

# row-stripe bg per category
CAT_COLORS = {
    "sunscreen":   ("FFF9C4", "FFF3A0"),
    "serum":       ("E8F5E9", "C8E6C9"),
    "moisturizer": ("E3F2FD", "BBDEFB"),
    "mask":        ("F3E5F5", "E1BEE7"),
    "cleanser":    ("FBE9E7", "FFCCBC"),
    "toner":       ("E0F7FA", "B2EBF2"),
    "eye_care":    ("FCE4EC", "F8BBD0"),
    "lip_care":    ("FFF8E1", "FFECB3"),
    "exfoliator":  ("E8EAF6", "C5CAE9"),
    "oil":         ("F9FBE7", "F0F4C3"),
    "unknown":     (WHITE, LIGHT_BG),
}

def row_bg(major_category, stripe):
    c1, c2 = CAT_COLORS.get(major_category, (WHITE, LIGHT_BG))
    return c2 if stripe else c1

COLUMNS = [
    ("#",             5),
    ("Product Name",  38),
    ("Brand",         18),
    ("Category",      15),
    ("Subtype",       22),
    ("Active Tags",   30),
    ("Functions",     30),
    ("URL",           45),
    ("Ingredients",   80),
]

TAG_BADGE = {
    "retinol":          ("FFD54F", "000000"),
    "retinal":          ("FF8A65", "FFFFFF"),
    "aha":              ("AED581", "000000"),
    "bha":              ("4DB6AC", "FFFFFF"),
    "pha":              ("80CBC4", "000000"),
    "vitamin_c":        ("FFB74D", "000000"),
    "niacinamide":      ("BA68C8", "FFFFFF"),
    "peptide":          ("4FC3F7", "000000"),
    "ceramide":         ("81C784", "000000"),
    "zinc_oxide":       ("90A4AE", "FFFFFF"),
    "titanium_dioxide": ("B0BEC5", "000000"),
}

def build_excel(products_data, filepath):
    wb = Workbook()

    # ─── Sheet 1: Products ───────────────────────────────────────
    ws = wb.active
    ws.title = "Products"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title bar (row 1)
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value     = "✨  Skincare Ingredient Database  ✨"
    title.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    title.fill      = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # Column headers (row 2)
    for ci, (label, _) in enumerate(COLUMNS, start=1):
        header_cell(ws, 2, ci, label, bg=TEAL, size=10)
    ws.row_dimensions[2].height = 28

    # Column widths
    for ci, (_, w) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    for ri, p in enumerate(products_data, start=1):
        excel_row = ri + 2
        stripe    = ri % 2 == 0
        bg        = row_bg(p["major_category"], stripe)

        ws.row_dimensions[excel_row].height = 22

        data_cell(ws, excel_row, 1, ri,    bg=bg, align="center", bold=True)
        data_cell(ws, excel_row, 2, p["name"],          bg=bg)
        data_cell(ws, excel_row, 3, p["brand"],         bg=bg)
        data_cell(ws, excel_row, 4, p["major_category"].upper(), bg=bg,
                    align="center", color=TEAL, bold=True)
        data_cell(ws, excel_row, 5, p["subtype"],       bg=bg)
        data_cell(ws, excel_row, 6, p["active_tags"],   bg=bg, wrap=True)
        data_cell(ws, excel_row, 7, p["function_tags"], bg=bg, wrap=True)

        # URL as hyperlink
        c_url = ws.cell(row=excel_row, column=8, value=p["product_url"])
        c_url.hyperlink   = p["product_url"]
        c_url.font        = Font(name="Arial", size=10, color="0563C1",
                                    underline="single")
        c_url.fill        = PatternFill("solid", fgColor=bg)
        c_url.alignment   = Alignment(horizontal="left", vertical="center")
        c_url.border      = make_border()

        data_cell(ws, excel_row, 9, p["ingredients_raw"], bg=bg, wrap=True)

    # AutoFilter on row 2
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(products_data)+2}"

    # ─── Sheet 2: Summary ────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:D1")
    t = ws2["A1"]
    t.value     = "Summary Statistics"
    t.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 14

    # Category counts
    header_cell(ws2, 2, 1, "Category",    bg=TEAL, size=10)
    header_cell(ws2, 2, 2, "Count",       bg=TEAL, size=10)
    header_cell(ws2, 2, 3, "Active Tag",  bg=TEAL, size=10)
    header_cell(ws2, 2, 4, "Count",       bg=TEAL, size=10)
    ws2.row_dimensions[2].height = 24

    from collections import Counter
    cat_counts = Counter(p["major_category"] for p in products_data)
    tag_counts = Counter(
        tag
        for p in products_data
        for tag in p["active_tags"].split(",") if tag
    )

    cats = sorted(cat_counts.items(), key=lambda x: -x[1])
    tags = sorted(tag_counts.items(), key=lambda x: -x[1])

    max_rows = max(len(cats), len(tags))
    for i in range(max_rows):
        r = i + 3
        bg = LIGHT_BG if i % 2 == 0 else WHITE
        if i < len(cats):
            data_cell(ws2, r, 1, cats[i][0], bg=bg)
            data_cell(ws2, r, 2, cats[i][1], bg=bg, align="center", bold=True)
        if i < len(tags):
            data_cell(ws2, r, 3, tags[i][0], bg=bg)
            data_cell(ws2, r, 4, tags[i][1], bg=bg, align="center", bold=True)
        ws2.row_dimensions[r].height = 20

    # Total row
    total_row = max_rows + 3
    header_cell(ws2, total_row, 1, "TOTAL", bg=NAVY, size=10)
    header_cell(ws2, total_row, 2, len(products_data), bg=NAVY, size=10)

    # ─── Sheet 3: Ingredient Index ───────────────────────────────
    ws3 = wb.create_sheet("Ingredient Index")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value     = "Ingredient Frequency Index"
    t3.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    t3.fill      = PatternFill("solid", fgColor=NAVY)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32
    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 18

    header_cell(ws3, 2, 1, "Ingredient",         bg=TEAL, size=10)
    header_cell(ws3, 2, 2, "# Products",         bg=TEAL, size=10)
    header_cell(ws3, 2, 3, "% Coverage",         bg=TEAL, size=10)
    ws3.row_dimensions[2].height = 24

    all_ingr = Counter()
    for p in products_data:
        for ingr in p["ingredients_list"].split(","):
            ingr = ingr.strip()
            if ingr:
                all_ingr[ingr] += 1

    total = len(products_data) or 1
    for i, (ingr, cnt) in enumerate(all_ingr.most_common(200)):
        r  = i + 3
        bg = LIGHT_BG if i % 2 == 0 else WHITE
        pct = cnt / total * 100
        data_cell(ws3, r, 1, ingr,            bg=bg)
        data_cell(ws3, r, 2, cnt,             bg=bg, align="center", bold=True)
        data_cell(ws3, r, 3, f"{pct:.1f}%",  bg=bg, align="center")
        ws3.row_dimensions[r].height = 18

    ws3.auto_filter.ref = f"A2:C{len(all_ingr.most_common(200))+2}"

    wb.save(filepath)
    print(f"📊 Excel saved → {filepath}")

# ===============================
# SETUP DRIVER
# ===============================

chrome_options = Options()
if HEADLESS:
    chrome_options.add_argument("--headless=new")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument(
    "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 Chrome/120 Safari/537.36"
)

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=chrome_options
)
wait = WebDriverWait(driver, 10)
products_data = []

# ===============================
# LOAD LINKS
# ===============================

if not os.path.exists(INPUT_FILE):
    raise FileNotFoundError(f"❌ {INPUT_FILE} not found")

with open(INPUT_FILE, "r", encoding="utf-8") as f:
    lines = [l.strip() for l in f if l.strip() and not l.startswith("#")]

print(f"📋 Loaded {len(lines)} links")

# ===============================
# SCRAPING LOOP
# ===============================

for idx, line in enumerate(lines):
    if "|" in line:
        link, major_category = line.split("|", 1)
        major_category = major_category.strip().lower()
    else:
        link = line.strip()
        major_category = None           # will be resolved after scraping

    print(f"[{idx+1}/{len(lines)}] {link[:70]}...")

    scraped = False
    for attempt in range(3):
        try:
            driver.get(link)
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "h1")))

            # Name
            try:
                name = clean_text(driver.find_element(By.TAG_NAME, "h1").text)
            except:
                name = ""

            # Brand
            try:
                brand = clean_text(driver.find_element(
                    By.CSS_SELECTOR, ".product-brand-title a").text)
            except:
                brand = ""

            # Ingredients
            try:
                els = driver.find_elements(By.CSS_SELECTOR, "a.ingred-link")
                ingredients_list = [clean_text(e.text) for e in els if e.text.strip()]
            except:
                ingredients_list = []

            ingredients_raw = clean_text(", ".join(ingredients_list))

            # ── Auto-detect category if not provided ──
            if not major_category or major_category == "unknown":
                major_category = guess_category(link, name, ingredients_raw)

            subtype        = detect_subtype(name, ingredients_raw, major_category)
            active_tags    = detect_tags(ingredients_raw, ACTIVE_RULES)
            function_tags  = detect_tags(ingredients_raw, FUNCTION_RULES)

            products_data.append({
                "product_url":      link,
                "name":             name,
                "brand":            brand,
                "major_category":   major_category,
                "subtype":          subtype,
                "active_tags":      ",".join(active_tags),
                "function_tags":    ",".join(function_tags),
                "ingredients_raw":  ingredients_raw,
                "ingredients_list": ",".join(ingredients_list),
            })

            scraped = True
            break

        except Exception as e:
            print(f"  ⚠️  Retry {attempt+1}: {e}")
            time.sleep(2 ** attempt)

    if not scraped:
        print(f"  ❌ Failed after 3 attempts: {link}")
        products_data.append({
            "product_url":      link,
            "name":             "SCRAPE_FAILED",
            "brand":            "",
            "major_category":   major_category or "unknown",
            "subtype":          "",
            "active_tags":      "",
            "function_tags":    "",
            "ingredients_raw":  "",
            "ingredients_list": "",
        })

    # Autosave (CSV backup)
    if (idx + 1) % AUTOSAVE_EVERY == 0:
        pd.DataFrame(products_data).to_csv(
            "autosave.csv", index=False, encoding="utf-8-sig")
        print("  💾 Autosaved CSV backup...")

    time.sleep(random.uniform(0.8, 1.8))

driver.quit()

# ===============================
# FINAL SAVE → EXCEL
# ===============================

build_excel(products_data, OUTPUT_FILE)

# Also save CSV as backup
csv_file = OUTPUT_FILE.replace(".xlsx", "_backup.csv")
pd.DataFrame(products_data).to_csv(csv_file, index=False, encoding="utf-8-sig")

print(f"\n✅ All done!")
print(f"   📊 Excel  → {OUTPUT_FILE}")
print(f"   📄 CSV    → {csv_file}")
print(f"   Total products scraped: {len(products_data)}")