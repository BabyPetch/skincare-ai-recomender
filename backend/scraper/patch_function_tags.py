"""
patch_function_tags.py
----------------------
Re-apply function_tags ทุกแถวด้วย rules ที่ครอบคลุมขึ้น

Usage:
    python patch_function_tags.py --input data_products/incidecoder_20260225_143616_patched.csv
"""

import argparse
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side


FUNCTION_RULES = {
    "brightening":    ["ascorb", "niacinamide", "arbutin", "kojic", "tranexamic",
                       "glutathione", "vitamin c", "licorice", "belides"],
    "anti_aging":     ["retinol", "retinal", "peptide", "bakuchiol", "ubiquinone",
                       "coenzyme q", "adenosine", "argireline", "matrixyl"],
    "acne_control":   ["salicylic acid", "benzoyl peroxide", "sulfur", "zinc",
                       "tea tree", "witch hazel"],
    "calming":        ["centella", "allantoin", "madecassoside", "asiaticoside",
                       "bisabolol", "aloe", "chamomile", "oat", "avena",
                       "panthenol", "licorice root", "dipotassium glycyrrhizate"],
    "barrier_repair": ["ceramide", "cholesterol", "fatty acid", "sphingolipid",
                       "glycosphingolipid", "phytosterol", "phospholipid"],
    "hydrating":      ["hyaluronic", "sodium hyaluronate", "glycerin", "glycerine",
                       "sorbitol", "urea", "sodium pca", "betaine",
                       "trehalose", "squalane", "sodium lactate"],
    "exfoliating":    ["glycolic acid", "lactic acid", "mandelic acid",
                       "salicylic acid", "gluconolactone", "citric acid",
                       "malic acid", "tartaric acid"],
    "antioxidant":    ["tocopherol", "vitamin e", "resveratrol", "ferulic",
                       "ubiquinone", "coenzyme q", "green tea", "camellia",
                       "ascorbyl", "superoxide dismutase"],
    "moisturizing":   ["petrolatum", "paraffin", "dimethicone", "shea butter",
                       "jojoba", "argan", "squalane", "caprylic", "triglyceride"],
}


def detect_function_tags(ingredients: str) -> str:
    ingr_lower = str(ingredients).lower()
    tags = []
    for tag, keywords in FUNCTION_RULES.items():
        if any(kw in ingr_lower for kw in keywords):
            tags.append(tag)
    return ",".join(sorted(tags)) if tags else ""


def _border():
    s = Side(style="thin", color="C8D8E4")
    return Border(left=s, right=s, top=s, bottom=s)


def patch_csv(input_path: Path, output_path: Path):
    df = pd.read_csv(input_path)
    print(f"  Loaded {len(df):,} rows")

    ingr_col = "ingredients_list" if "ingredients_list" in df.columns else "ingredients_raw"

    before_null = df["function_tags"].isna().sum()
    df["function_tags"] = df[ingr_col].apply(detect_function_tags)
    after_null = (df["function_tags"] == "").sum()

    print(f"  function_tags null before : {before_null}")
    print(f"  function_tags empty after : {after_null}")

    df.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"  CSV saved -> {output_path}")
    return df


def patch_excel(input_path: Path, output_path: Path, df: pd.DataFrame):
    wb = load_workbook(input_path)
    ws = wb["Products"] if "Products" in wb.sheetnames else wb.active

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if str(cell.value).lower() in ("name", "product name"):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        print("  WARNING: หา header row ไม่เจอใน Excel")
        return

    headers = {str(ws.cell(row=header_row, column=c).value).lower(): c
               for c in range(1, ws.max_column + 1)}

    ftag_col = headers.get("function_tags") or headers.get("functions")
    url_col  = headers.get("product_url") or headers.get("url")

    if not ftag_col:
        print("  WARNING: ไม่พบ function_tags column ใน Excel")
        return

    if url_col and "product_url" in df.columns:
        url_to_ftag = dict(zip(df["product_url"], df["function_tags"]))
        for r in range(header_row + 1, ws.max_row + 1):
            url_val  = ws.cell(row=r, column=url_col).value
            ftag_val = url_to_ftag.get(str(url_val).strip(), "")
            cell = ws.cell(row=r, column=ftag_col, value=ftag_val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border    = _border()

    wb.save(output_path)
    print(f"  Excel saved -> {output_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  required=True)
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    input_path = Path(args.input)
    out_dir    = Path(args.output).parent if args.output else input_path.parent
    base       = Path(args.output).stem   if args.output else input_path.stem

    csv_out  = out_dir / f"{base}.csv"
    xlsx_out = out_dir / f"{base}.xlsx"

    print("=" * 50)
    print("  Patch Function Tags")
    print("=" * 50)

    print(f"\nPatching CSV: {input_path.name}")
    df = patch_csv(input_path, csv_out)

    xlsx_same = input_path.with_suffix(".xlsx")
    if xlsx_same.exists():
        print(f"\nพบ Excel: {xlsx_same.name} — patch ด้วยเลย")
        patch_excel(xlsx_same, xlsx_out, df)

    # Stats
    print(f"\nFunction tags distribution:")
    exploded = df["function_tags"].str.split(",").explode()
    print(exploded[exploded != ""].value_counts().to_string())
    print(f"\nDone!")


if __name__ == "__main__":
    main()