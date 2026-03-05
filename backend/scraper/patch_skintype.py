"""
patch_skintype.py
-----------------
เพิ่ม column 'skintype' ให้ CSV และ Excel เดิมพร้อมกัน
ไม่ต้อง scrape ใหม่

Usage:
    python patch_skintype.py --input incidecoder_XXXXXX.csv
    python patch_skintype.py --input incidecoder_XXXXXX.xlsx
    python patch_skintype.py --input incidecoder_XXXXXX.csv --output patched.csv
"""

import argparse
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ================================================================
# SKINTYPE RULES
# ================================================================

SKINTYPE_RULES = {
    "oily": [
        "salicylic acid", "niacinamide", "zinc oxide", "kaolin", "bentonite",
        "clay", "witch hazel", "tea tree", "sulfur", "zinc pca",
    ],
    "dry": [
        "ceramide", "shea butter", "squalane", "hyaluronic acid",
        "sodium hyaluronate", "glycerin", "glycerine", "petrolatum",
        "lanolin", "urea", "cholesterol", "fatty acid",
    ],
    "sensitive": [
        "centella asiatica", "madecassoside", "madecassic acid",
        "asiaticoside", "allantoin", "bisabolol", "aloe barbadensis",
        "aloe vera", "panthenol", "oat", "avena sativa",
        "chamomile", "licorice root", "dipotassium glycyrrhizate",
    ],
    "combination": [
        "glycolic acid", "lactic acid", "mandelic acid",
        "azelaic acid", "retinol", "retinal",
    ],
}

FUNCTION_TAG_HINTS = {
    "acne_control":   "oily",
    "exfoliating":    "combination",
    "barrier_repair": "dry",
    "calming":        "sensitive",
    "hydrating":      "dry",
    "brightening":    "combination",
    "anti_aging":     "combination",
}

# ================================================================
# DETECT SKINTYPE
# ================================================================

def detect_skintype(ingredients: str, function_tags: str) -> str:
    ingr_lower = str(ingredients).lower()
    tags_lower = str(function_tags).lower()

    scores = {st: 0 for st in SKINTYPE_RULES}
    for skintype, keywords in SKINTYPE_RULES.items():
        for kw in keywords:
            if kw in ingr_lower:
                scores[skintype] += 1

    max_score = max(scores.values())
    if max_score > 0:
        winners = [st for st, sc in scores.items() if sc == max_score]
        return ",".join(winners)

    for tag, skintype in FUNCTION_TAG_HINTS.items():
        if tag in tags_lower:
            return skintype

    return "all"

# ================================================================
# PATCH CSV
# ================================================================

def patch_csv(input_path: Path, output_path: Path):
    df = pd.read_csv(input_path)
    print(f"  Loaded {len(df):,} rows")

    ingr_col = "ingredients_list" if "ingredients_list" in df.columns else "ingredients_raw"
    ftag_col = "function_tags" if "function_tags" in df.columns else None

    df["skintype"] = df.apply(
        lambda row: detect_skintype(
            row.get(ingr_col, ""),
            row.get(ftag_col, "") if ftag_col else ""
        ), axis=1
    )

    df.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"  CSV saved -> {output_path}")
    return df

# ================================================================
# PATCH EXCEL
# ================================================================

def _border():
    s = Side(style="thin", color="C8D8E4")
    return Border(left=s, right=s, top=s, bottom=s)

def patch_excel(input_path: Path, output_path: Path, df: pd.DataFrame):
    wb = load_workbook(input_path)
    ws = wb["Products"] if "Products" in wb.sheetnames else wb.active

    # หา header row
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if str(cell.value).lower() in ("name", "product name"):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        print("  WARNING: หา header row ไม่เจอใน Excel จะข้าม Excel ไป")
        return

    headers = {str(ws.cell(row=header_row, column=c).value).lower(): c
               for c in range(1, ws.max_column + 1)}

    if "skintype" in headers:
        skintype_col = headers["skintype"]
        print(f"  พบ 'skintype' column อยู่แล้ว (col {skintype_col}) จะ overwrite")
    else:
        skintype_col = ws.max_column + 1
        cell = ws.cell(row=header_row, column=skintype_col, value="skintype")
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D7D9A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
        print(f"  เพิ่ม 'skintype' column ที่ col {skintype_col}")

    url_col = headers.get("product_url") or headers.get("url")

    if url_col and "product_url" in df.columns:
        url_to_skintype = dict(zip(df["product_url"], df["skintype"]))
        for r in range(header_row + 1, ws.max_row + 1):
            url_val = ws.cell(row=r, column=url_col).value
            skintype_val = url_to_skintype.get(str(url_val).strip(), "all")
            cell = ws.cell(row=r, column=skintype_col, value=skintype_val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()
    else:
        ingr_col = headers.get("ingredients_list") or headers.get("ingredients_raw")
        ftag_col = headers.get("function_tags")
        for r in range(header_row + 1, ws.max_row + 1):
            ingr = ws.cell(row=r, column=ingr_col).value if ingr_col else ""
            ftag = ws.cell(row=r, column=ftag_col).value if ftag_col else ""
            skintype_val = detect_skintype(ingr or "", ftag or "")
            cell = ws.cell(row=r, column=skintype_col, value=skintype_val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()

    wb.save(output_path)
    print(f"  Excel saved -> {output_path}")

# ================================================================
# MAIN
# ================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  required=True, help="path to original CSV or Excel")
    parser.add_argument("--output", default=None,  help="output path (optional)")
    args = parser.parse_args()

    input_path = Path(args.input)
    ext = input_path.suffix.lower()

    if args.output:
        base_output = Path(args.output).stem
        out_dir = Path(args.output).parent
    else:
        base_output = input_path.stem + "_patched"
        out_dir = input_path.parent

    csv_out  = out_dir / f"{base_output}.csv"
    xlsx_out = out_dir / f"{base_output}.xlsx"

    print("=" * 50)
    print("  Patch Skintype")
    print("=" * 50)

    if ext == ".csv":
        print(f"\nPatching CSV: {input_path.name}")
        df = patch_csv(input_path, csv_out)

        # ถ้ามี xlsx ชื่อเดียวกันอยู่ด้วย patch ด้วยเลย
        xlsx_same = input_path.with_suffix(".xlsx")
        if xlsx_same.exists():
            print(f"\nพบ Excel ชื่อเดียวกัน: {xlsx_same.name} — patch ด้วยเลย")
            patch_excel(xlsx_same, xlsx_out, df)
        else:
            print(f"\nไม่พบ Excel ชื่อเดียวกัน (ถ้ามีให้วางไว้ที่เดียวกับ CSV จะ patch อัตโนมัติ)")

    elif ext in (".xlsx", ".xls"):
        print(f"\nReading Excel: {input_path.name}")
        df_raw = pd.read_excel(input_path, sheet_name=0, header=1)
        df_raw.columns = [str(c).strip().lower().replace(" ", "_") for c in df_raw.columns]

        ingr_col = "ingredients_list" if "ingredients_list" in df_raw.columns else "ingredients_raw"
        ftag_col = "function_tags" if "function_tags" in df_raw.columns else None

        df_raw["skintype"] = df_raw.apply(
            lambda row: detect_skintype(
                row.get(ingr_col, ""),
                row.get(ftag_col, "") if ftag_col else ""
            ), axis=1
        )

        df_raw.to_csv(csv_out, index=False, encoding="utf-8-sig")
        print(f"  CSV saved -> {csv_out}")

        patch_excel(input_path, xlsx_out, df_raw)
        df = df_raw

    else:
        print(f"ไม่รองรับไฟล์นามสกุล {ext}")
        return

    # Stats
    print(f"\nSkintype distribution:")
    exploded = df["skintype"].str.split(",").explode()
    print(exploded.value_counts().to_string())
    empty = (df["skintype"] == "all").sum()
    print(f"\n  'all' (ไม่ match rule ใดเลย): {empty:,} / {len(df):,} rows")
    print("\nDone!")

if __name__ == "__main__":
    main()