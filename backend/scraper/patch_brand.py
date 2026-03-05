"""
patch_brand.py
--------------
Extract brand จากชื่อสินค้า แล้ว patch ลง CSV + Excel

Usage:
    python patch_brand.py --input data_products/incidecoder_20260225_143616_patched.csv
"""

import argparse
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _border():
    s = Side(style="thin", color="C8D8E4")
    return Border(left=s, right=s, top=s, bottom=s)


BRAND_OVERRIDES = {
    "the mask dr.": "the mask dr.",
}

def extract_brand(name: str) -> str:
    """ตัดคำสุดท้าย (category) ออก เหลือเป็น brand"""
    name_str = str(name)
    for key, brand in BRAND_OVERRIDES.items():
        if key.lower() in name_str.lower():
            return brand
    parts = name_str.rsplit(' ', 1)
    return parts[0].strip() if len(parts) > 1 else name_str.strip()


def patch_csv(input_path: Path, output_path: Path):
    df = pd.read_csv(input_path)
    print(f"  Loaded {len(df):,} rows")

    df['brand'] = df['name'].apply(extract_brand)

    df.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"  CSV saved -> {output_path}")
    return df


def patch_excel(input_path: Path, output_path: Path, df: pd.DataFrame):
    wb = load_workbook(input_path)
    ws = wb["Products"] if "Products" in wb.sheetnames else wb.active

    # หา header row
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if str(cell.value).lower() in ("name", "product name"):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        print("  WARNING: หา header row ไม่เจอใน Excel")
        return

    headers = {str(ws.cell(row=header_row, column=c).value).lower(): c
               for c in range(1, ws.max_column + 1)}

    brand_col = headers.get("brand")
    url_col   = headers.get("product_url") or headers.get("url")

    if not brand_col:
        print("  WARNING: ไม่พบ brand column ใน Excel")
        return

    if url_col and "product_url" in df.columns:
        url_to_brand = dict(zip(df["product_url"], df["brand"]))
        for r in range(header_row + 1, ws.max_row + 1):
            url_val   = ws.cell(row=r, column=url_col).value
            brand_val = url_to_brand.get(str(url_val).strip(), "")
            cell = ws.cell(row=r, column=brand_col, value=brand_val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = _border()
    else:
        name_col = headers.get("name") or headers.get("product name")
        for r in range(header_row + 1, ws.max_row + 1):
            name_val  = ws.cell(row=r, column=name_col).value if name_col else ""
            brand_val = extract_brand(name_val or "")
            cell = ws.cell(row=r, column=brand_col, value=brand_val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = _border()

    wb.save(output_path)
    print(f"  Excel saved -> {output_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  required=True)
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    input_path = Path(args.input)

    if args.output:
        out_dir     = Path(args.output).parent
        base_output = Path(args.output).stem
    else:
        out_dir     = input_path.parent
        base_output = input_path.stem  # overwrite ไฟล์เดิมเลย

    csv_out  = out_dir / f"{base_output}.csv"
    xlsx_out = out_dir / f"{base_output}.xlsx"

    print("=" * 50)
    print("  Patch Brand")
    print("=" * 50)

    print(f"\nPatching CSV: {input_path.name}")
    df = patch_csv(input_path, csv_out)

    xlsx_same = input_path.with_suffix(".xlsx")
    if xlsx_same.exists():
        print(f"\nพบ Excel: {xlsx_same.name} — patch ด้วยเลย")
        patch_excel(xlsx_same, xlsx_out, df)

    # Stats
    print(f"\nTop 10 brands:")
    print(df['brand'].value_counts().head(10).to_string())
    print(f"\nDone! {len(df):,} rows patched")


if __name__ == "__main__":
    main()