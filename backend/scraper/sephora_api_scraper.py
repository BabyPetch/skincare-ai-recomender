"""
╔══════════════════════════════════════════════════════════════════╗
║   Sephora API Scraper  —  ใช้ browseSearchProduct API           ║
║   ไม่ต้องใช้ Selenium เลย — เรียก API โดยตรง                   ║
║                                                                  ║
║   pip install requests pandas openpyxl pillow                   ║
║   python sephora_api_scraper.py                                  ║
╚══════════════════════════════════════════════════════════════════╝
"""

import time, datetime, re, hashlib, json
from io import BytesIO
from pathlib import Path
from collections import Counter

import requests
import pandas as pd
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ================================================================
# CONFIG
# ================================================================

API_KEY    = "nQc7BFt78yJBvfYDKtle9APd5RrX984i"
BASE_URL   = "https://api-developer.sephora.com/v1/browseSearchProduct"
PAGE_SIZE  = 60
MAX_PAGES  = 5           # หน้าสูงสุดต่อ category (None = ทั้งหมด)
DELAY      = 1.0         # วินาทีระหว่าง request
IMG_SIZE   = (120, 120)
AUTOSAVE   = 50

OUTPUT_DIR = Path(r"C:\Users\User\Documents\GitHub\skincare-ai-recomender\backend\scraper\data_products")
IMG_DIR    = OUTPUT_DIR / "images"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
IMG_DIR.mkdir(parents=True, exist_ok=True)

timestamp   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_XLSX = OUTPUT_DIR / f"sephora_{timestamp}.xlsx"
OUTPUT_CSV  = OUTPUT_DIR / f"sephora_{timestamp}.csv"

# Sephora category IDs (จาก URL patterns)
CATEGORIES = {
    "moisturizer": {"categoryId": "C_moisturizer_skincare", "display": "moisturizer"},
    "serum":       {"categoryId": "C_face-serum",            "display": "serum"},
    "sunscreen":   {"categoryId": "C_sunscreen-sun-protection", "display": "sunscreen"},
    "cleanser":    {"categoryId": "C_face-wash-facial-cleanser", "display": "cleanser"},
    "eye_care":    {"categoryId": "C_eye-cream",             "display": "eye_care"},
    "mask":        {"categoryId": "C_face-mask-skincare",    "display": "mask"},
    "toner":       {"categoryId": "C_facial-toners-toners",  "display": "toner"},
}

HEADERS = {
    "apikey":       API_KEY,
    "x-api-key":    API_KEY,
    "Accept":       "application/json",
    "Content-Type": "application/json",
    "User-Agent":   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Origin":       "https://www.sephora.com",
    "Referer":      "https://www.sephora.com/",
}

# ── Tag rules ────────────────────────────────────────────────────
ACTIVE_RULES = {
    "retinol":["retinol"],"retinal":["retinal"],
    "aha":["glycolic acid","lactic acid","mandelic acid"],
    "bha":["salicylic acid"],"pha":["gluconolactone"],
    "vitamin_c":["ascorb","ascorbyl"],"niacinamide":["niacinamide"],
    "peptide":["peptide"],"ceramide":["ceramide"],
    "zinc_oxide":["zinc oxide"],"titanium_dioxide":["titanium dioxide"],
}
FUNCTION_RULES = {
    "brightening":["ascorb","niacinamide","arbutin"],
    "anti_aging":["retinol","retinal","peptide"],
    "acne_control":["salicylic acid"],"calming":["centella","allantoin"],
    "barrier_repair":["ceramide"],"hydrating":["hyaluronic","glycerin"],
    "exfoliating":["glycolic","salicylic","lactic"],
}

def clean(t): return " ".join(str(t).split()).strip()
def detect_tags(text, rules):
    t = text.lower()
    return sorted(tag for tag, kws in rules.items() if any(k in t for k in kws))
def detect_subtype(name, ingr, cat):
    t = (name+" "+ingr).lower()
    if cat=="sunscreen":
        m="zinc oxide" in t or "titanium dioxide" in t
        c=any(k in t for k in ["avobenzone","octinoxate","octocrylene"])
        return "hybrid_sunscreen" if m and c else ("mineral_sunscreen" if m else "chemical_sunscreen")
    if "retinal" in t: return "retinal_serum"
    if "retinol" in t: return "retinol_serum"
    if "salicylic acid" in t: return "bha_exfoliator"
    if "glycolic acid" in t: return "aha_exfoliator"
    return cat

# ================================================================
# IMAGE DOWNLOAD
# ================================================================

def download_image(img_url: str) -> Path | None:
    if not img_url or img_url.startswith("data:"): return None
    try:
        h = hashlib.md5(img_url.encode()).hexdigest()[:12]
        p = IMG_DIR / f"{h}.jpg"
        if p.exists(): return p
        r = requests.get(img_url, headers={"User-Agent": HEADERS["User-Agent"],
                                            "Referer": "https://www.sephora.com/"},
                         timeout=10)
        if r.status_code != 200: return None
        img = PILImage.open(BytesIO(r.content)).convert("RGB")
        img.thumbnail(IMG_SIZE, PILImage.LANCZOS)
        img.save(p, "JPEG", quality=85)
        return p
    except: return None

# ================================================================
# API CALL — browseSearchProduct
# ================================================================

def fetch_category_page(category_id: str, page: int) -> dict | None:
    """
    เรียก Sephora browseSearchProduct API
    return JSON หรือ None ถ้า error
    """
    params = {
        "apikey":       API_KEY,
        "categoryId":   category_id,
        "currentPage":  page,
        "pageSize":     PAGE_SIZE,
        "sortBy":       "TOP_SELLERS",
        "loc":          "en-US",
        "country":      "US",
        "channel":      "desktop",
        "includeRegionsMap": "true",
    }
    try:
        resp = requests.get(BASE_URL, headers=HEADERS, params=params, timeout=15)
        print(f"     HTTP {resp.status_code}", end="")

        if resp.status_code == 200:
            print(f" ✅")
            return resp.json()
        elif resp.status_code == 401:
            print(f" ❌ Unauthorized — API key หมดอายุ")
            return None
        elif resp.status_code == 403:
            print(f" ❌ Forbidden")
            return None
        elif resp.status_code == 404:
            print(f" ❌ Category not found")
            return None
        else:
            print(f" ⚠️  {resp.text[:200]}")
            # ลอง save response เพื่อ debug
            Path(f"debug_response_{resp.status_code}.json").write_text(resp.text[:2000])
            return None
    except Exception as e:
        print(f" ❌ {e}")
        return None


def parse_products(data: dict, category: str) -> list[dict]:
    """แกะ JSON response → list of product dicts"""
    products = []

    # Sephora API อาจส่ง products ใน path ต่างๆ
    raw_list = (
        data.get("products") or
        data.get("data", {}).get("products") or
        data.get("skus") or
        data.get("result", {}).get("products") or
        data.get("items") or
        []
    )

    # Debug: print structure ถ้าว่าง
    if not raw_list:
        print(f"\n  ⚠️  ไม่เจอ products key — keys ที่มี: {list(data.keys())[:10]}")
        # Save full response เพื่อ debug
        Path(f"debug_{category}.json").write_text(json.dumps(data, indent=2)[:3000])
        print(f"     บันทึก debug_{category}.json แล้ว")
        return []

    for item in raw_list:
        try:
            # ดึงข้อมูลพื้นฐาน — Sephora มีหลาย schema
            name  = clean(item.get("displayName") or item.get("name") or item.get("productName") or "")
            brand = clean(item.get("brandName") or item.get("brand", {}).get("displayName") or "")
            sku   = item.get("skuId") or item.get("currentSku", {}).get("skuId") or ""
            pid   = item.get("productId") or item.get("id") or ""

            # Price
            price_data = item.get("currentSku", {}).get("listPrice") or item.get("listPrice") or ""
            if isinstance(price_data, dict):
                price = price_data.get("value") or price_data.get("formatted") or ""
            else:
                price = str(price_data)

            # Rating
            rating      = str(item.get("rating") or item.get("reviews", {}).get("rating") or "")
            rating_count= str(item.get("reviews") or item.get("reviewCount") or "")
            if isinstance(item.get("reviews"), dict):
                rating       = str(item["reviews"].get("rating",""))
                rating_count = str(item["reviews"].get("currentSyndicationSource","") or
                                   item["reviews"].get("totalReviews",""))

            # Image
            img_url = (
                item.get("heroImage") or
                item.get("currentSku", {}).get("heroImage") or
                item.get("imageUrl") or
                item.get("image", {}).get("href") or
                ""
            )
            # เติม domain ถ้าเป็น relative path
            if img_url and img_url.startswith("//"):
                img_url = "https:" + img_url
            elif img_url and img_url.startswith("/"):
                img_url = "https://www.sephora.com" + img_url

            # URL
            product_url = item.get("targetUrl") or item.get("url") or ""
            if product_url and not product_url.startswith("http"):
                product_url = "https://www.sephora.com" + product_url

            # Ingredients (บางครั้งอยู่ใน SKU details)
            ingr_raw = clean(
                item.get("ingredients") or
                item.get("currentSku", {}).get("ingredients") or
                ""
            )
            ingr_list = [i.strip() for i in re.split(r",\s*", ingr_raw) if i.strip()] if ingr_raw else []

            # Download image
            img_local = download_image(img_url)

            subtype = detect_subtype(name, ingr_raw, category)

            products.append({
                "source":           "Sephora",
                "product_url":      product_url,
                "name":             name,
                "brand":            brand,
                "sku_id":           sku,
                "product_id":       pid,
                "major_category":   category,
                "subtype":          subtype,
                "price":            price,
                "rating":           rating,
                "rating_count":     rating_count,
                "active_tags":      ",".join(detect_tags(ingr_raw, ACTIVE_RULES)),
                "function_tags":    ",".join(detect_tags(ingr_raw, FUNCTION_RULES)),
                "ingredients_raw":  ingr_raw,
                "ingredients_list": ",".join(ingr_list),
                "image_url":        img_url,
                "image_local":      str(img_local) if img_local else "",
            })
        except Exception as e:
            print(f"    ⚠️  Parse error: {e}")
            continue

    return products


def get_total_pages(data: dict) -> int:
    """หาจำนวนหน้าทั้งหมดจาก response"""
    total = (
        data.get("totalProducts") or
        data.get("total") or
        data.get("data", {}).get("total") or
        data.get("pagination", {}).get("total") or
        0
    )
    try:
        return max(1, -(-int(total) // PAGE_SIZE))  # ceiling division
    except:
        return 1

# ================================================================
# EXCEL BUILDER  (same style as INCIDecoder)
# ================================================================

NAVY="1B3A5C"; TEAL="2D7D9A"; WHITE="FFFFFF"; LIGHT="EBF4F8"
PINK="C75B7A"; PINK_LIGHT="FFF0F5"
CAT_COLORS={
    "moisturizer":("E3F2FD","BBDEFB"),"serum":("E8F5E9","C8E6C9"),
    "sunscreen":("FFF9C4","FFF3A0"),"toner":("E0F7FA","B2EBF2"),
    "cleanser":("FBE9E7","FFCCBC"),"eye_care":("FCE4EC","F8BBD0"),
    "mask":("F3E5F5","E1BEE7"),"exfoliator":("E8EAF6","C5CAE9"),
}

def _b():
    s=Side(style="thin",color="C8D8E4")
    return Border(left=s,right=s,top=s,bottom=s)
def _h(ws,r,c,v,bg=NAVY,fg=WHITE,sz=10):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(name="Arial",bold=True,size=sz,color=fg)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=_b()
def _d(ws,r,c,v,bg=WHITE,bold=False,wrap=False,align="left",color="1A1A1A"):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(name="Arial",size=10,bold=bold,color=color)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    cell.border=_b()

COLUMNS=[
    ("Image",16,"_image"), ("#",5,"_num"), ("Source",10,"source"),
    ("Brand",18,"brand"), ("Product Name",40,"name"),
    ("Category",14,"major_category"), ("Subtype",22,"subtype"),
    ("Price",10,"price"), ("Rating",10,"rating"), ("Reviews",10,"rating_count"),
    ("Active Tags",30,"active_tags"), ("Functions",28,"function_tags"),
    ("URL",42,"product_url"), ("Image URL",42,"image_url"),
    ("Ingredients",90,"ingredients_raw"),
]

def build_excel(products, filepath):
    print(f"\n📊 Building Excel …")
    wb=Workbook()
    ws=wb.active; ws.title="Sephora Products"
    ws.sheet_view.showGridLines=False; ws.freeze_panes="B3"

    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    t=ws["A1"]
    t.value=f"💄  Sephora Best Sellers  —  {len(products):,} products  ✨"
    t.font=Font(name="Arial",bold=True,size=13,color=WHITE)
    t.fill=PatternFill("solid",fgColor=PINK)
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=32

    for ci,(label,width,_) in enumerate(COLUMNS,1):
        _h(ws,2,ci,label,bg=PINK)
        ws.column_dimensions[get_column_letter(ci)].width=width
    ws.row_dimensions[2].height=28

    for ri,p in enumerate(products,1):
        er=ri+2
        cat=p.get("major_category","unknown")
        c1,c2=CAT_COLORS.get(cat,(WHITE,LIGHT))
        bg=c2 if ri%2==0 else c1
        ws.row_dimensions[er].height=90

        for ci,(_,_,key) in enumerate(COLUMNS,1):
            cl=get_column_letter(ci)
            if key=="_image":
                il=p.get("image_local","")
                if il and Path(il).exists():
                    try:
                        xi=XLImage(il); xi.width=88; xi.height=88
                        ws.add_image(xi,f"{cl}{er}")
                    except: _d(ws,er,ci,"📷",bg=bg,align="center")
                else: _d(ws,er,ci,"",bg=bg)
            elif key=="_num":
                _d(ws,er,ci,ri,bg=bg,align="center",bold=True)
            elif key=="source":
                _d(ws,er,ci,"Sephora",bg=bg,align="center",bold=True,color=PINK)
            elif key=="major_category":
                _d(ws,er,ci,str(p.get(key,"")).upper(),bg=bg,align="center",bold=True,color=TEAL)
            elif key in("product_url","image_url"):
                val=p.get(key,"")
                cell=ws.cell(row=er,column=ci,value=val)
                if val: cell.hyperlink=val
                cell.font=Font(name="Arial",size=9,color="0563C1",underline="single")
                cell.fill=PatternFill("solid",fgColor=bg)
                cell.alignment=Alignment(horizontal="left",vertical="center")
                cell.border=_b()
            elif key=="ingredients_raw":
                _d(ws,er,ci,p.get(key,""),bg=bg,wrap=True)
            else:
                _d(ws,er,ci,p.get(key,""),bg=bg)

    ws.auto_filter.ref=f"A2:{get_column_letter(len(COLUMNS))}{len(products)+2}"

    # Summary sheet
    ws2=wb.create_sheet("Summary"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:D1"); t2=ws2["A1"]
    t2.value="Summary Statistics"
    t2.font=Font(name="Arial",bold=True,size=13,color=WHITE)
    t2.fill=PatternFill("solid",fgColor=PINK)
    t2.alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[1].height=30
    for ci,lbl in enumerate(["Category","Count","Active Tag","Count"],1):
        _h(ws2,2,ci,lbl,bg=PINK)
        ws2.column_dimensions[get_column_letter(ci)].width=22

    cat_c=Counter(p.get("major_category","?") for p in products)
    tag_c=Counter(t.strip() for p in products for t in p.get("active_tags","").split(",") if t.strip())
    cats=sorted(cat_c.items(),key=lambda x:-x[1])
    tags=sorted(tag_c.items(),key=lambda x:-x[1])
    for i in range(max(len(cats),len(tags))):
        r=i+3; bg=LIGHT if i%2==0 else WHITE
        if i<len(cats): _d(ws2,r,1,cats[i][0],bg=bg); _d(ws2,r,2,cats[i][1],bg=bg,align="center",bold=True)
        if i<len(tags): _d(ws2,r,3,tags[i][0],bg=bg); _d(ws2,r,4,tags[i][1],bg=bg,align="center",bold=True)
        ws2.row_dimensions[r].height=20
    tr=max(len(cats),len(tags))+3
    _h(ws2,tr,1,"TOTAL",bg=PINK); _h(ws2,tr,2,len(products),bg=PINK)

    wb.save(filepath)
    print(f"✅ Excel saved → {filepath}")

# ================================================================
# MAIN
# ================================================================

def run():
    print("="*60)
    print("  💄  Sephora API Scraper  —  No Selenium!")
    print(f"  📁  {OUTPUT_DIR}")
    print("="*60)

    all_products = []

    for cat_key, cat_info in CATEGORIES.items():
        cat_id  = cat_info["categoryId"]
        cat_name= cat_info["display"]

        print(f"\n{'='*50}")
        print(f"📂  {cat_name.upper()}")
        print(f"{'='*50}")

        page       = 1
        total_pages= 1
        cat_products = []

        while page <= total_pages:
            print(f"  📄 Page {page}/{total_pages} …", end="", flush=True)

            data = fetch_category_page(cat_id, page)

            if data is None:
                print(f"  ❌ Failed — skipping rest of {cat_name}")
                break

            # อัปเดต total pages จาก response แรก
            if page == 1:
                total_pages = min(get_total_pages(data), MAX_PAGES or 9999)
                print(f"     Total pages: {total_pages}")

            products = parse_products(data, cat_name)
            cat_products.extend(products)
            print(f"     +{len(products)} products  (cat total: {len(cat_products)})")

            if not products:
                print(f"  ⚠️  Empty page — stopping")
                break

            page += 1
            time.sleep(DELAY)

        all_products.extend(cat_products)
        print(f"  ✅ {cat_name}: {len(cat_products)} products")

        # Autosave
        if all_products and len(all_products) % AUTOSAVE < len(cat_products):
            pd.DataFrame(all_products).to_csv(
                OUTPUT_DIR/"autosave_sephora.csv", index=False, encoding="utf-8-sig")
            print(f"  💾 Autosaved {len(all_products)} total")

    # ── Deduplicate ──────────────────────────────────────────────
    seen, deduped = set(), []
    for p in all_products:
        key=(p.get("name","").lower(), p.get("brand","").lower())
        if key not in seen:
            seen.add(key); deduped.append(p)
    print(f"\n🧹 {len(all_products)} → {len(deduped)} unique products")

    if not deduped:
        print("\n❌ No data — API อาจเปลี่ยน endpoint หรือ key หมดอายุ")
        print("   ลองรัน sephora_capture.py ใหม่เพื่อดู endpoint ล่าสุด")
        return

    # ── Save ─────────────────────────────────────────────────────
    df = pd.DataFrame(deduped)
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    print(f"  📄 CSV  → {OUTPUT_CSV}")

    build_excel(deduped, OUTPUT_XLSX)

    img_ok = sum(1 for p in deduped if p.get("image_local"))
    print(f"\n{'='*60}")
    print(f"✅  DONE")
    print(f"   Products  : {len(deduped):,}")
    print(f"   Images    : {img_ok:,} / {len(deduped):,}")
    print(f"   📊 Excel  → {OUTPUT_XLSX}")
    print(f"   📄 CSV    → {OUTPUT_CSV}")
    print(f"{'='*60}")

if __name__ == "__main__":
    run()